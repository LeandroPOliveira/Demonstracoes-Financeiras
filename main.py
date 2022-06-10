import openpyxl
import pandas as pd
from datetime import datetime

wb = openpyxl.load_workbook('DF2.xlsx')  # Abrir arquivo com a base
sheets = wb.sheetnames
ws = wb['balancetes']

data_usuario = '31/03/2022'

data_cur = datetime.strptime(data_usuario, '%d/%m/%Y')

pos_coluna = 0
# Caso seja o primeiro trimestre, alterar ordem das colunas para excluir dados do ano anterior
# if data_cur.month == 1:
#
#     colunas = [['T:T', 'AF:AF'], ['H:H', 'T:T'], ['I:I', 'U:U'], ['J:J', 'V:V'], ['K:K', 'W:W'], ['L:L', 'X:X'],
#                ['M:M', 'Y:Y'], ['N:N', 'Z:Z'], ['O:O', 'AA:AA'], ['P:P', 'AB:AB'], ['Q:Q', 'AC:AC'],
#                ['R:R', 'AD:AD'], ['S:S', 'AE:AE']]
#
#     for col in colunas:
#         for src, dst in zip(ws[col[0]], ws[col[1]]):
#             dst.value = src.value
#
#         for row in ws[col[0][0] + '6:' + col[0][2] + '736']:
#             for cell in row:
#                 cell.value = None
#
#     for row in ws.iter_rows(min_row=4, min_col=8, max_col=19, max_row=4):
#         for index, cell in enumerate(reversed(row)):
#             # print(cell.value.year)
#             if cell.value.year != data_cur.year:
#                 cell.value = (datetime.strptime(f'{calendar.monthrange(data_cur.year, index + 1)[1]}/'
#                                                 f'{"0" + str(index + 1) if index + 1 < 10 else index + 1}/'
#                                                 f'{data_cur.year}', "%d/%m/%Y"))
#                 print(cell.value)
#
#     for row in ws.iter_rows(min_row=4, min_col=8, max_col=19, max_row=4):
#         for index, cell in enumerate(row):
#             if cell.value.month == data_cur.month:
#                 pos_coluna = index
#
#     pos_coluna += 8
#     print(pos_coluna)
#
# else:

for row in ws.iter_rows(min_row=4, min_col=8, max_col=19, max_row=4):
    for index, cell in enumerate(row):
        if cell.value.month == data_cur.month:
            pos_coluna = index

pos_coluna += 8
print(pos_coluna)
# ============================== Buscar dados de contas do razão =====================#
razao = pd.read_excel(f'razao{data_cur.month}.xlsx')
razao = pd.DataFrame(razao)
razao.fillna('', inplace=True)
razao['Texto'] = razao['Texto'].str.lower()

ppr, ppr1, ppr2, outras_rec, outras_desp, ir_dif, csll_dif, indebito, prov_ind, deliberacao, termok, vag, vag2 = \
    0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0

for index, row in razao.iterrows():
    if row['Atribuição'] == '615.03.1.1.01.01' and ('ppr' in row['Texto'] or 'partic' in row['Texto']):
        datas_trim = razao.loc[index, 'Montante em moeda interna']
        ppr += datas_trim
    elif row['Atribuição'] == '616.00.0.1.01.01' and ('ppr' in row['Texto'] or 'partic' in row['Texto']):
        lista1 = razao.loc[index, 'Montante em moeda interna']
        ppr1 += lista1
    elif row['Atribuição'] == '616.00.0.3.01.01' and ('ppr' in row['Texto'] or 'partic' in row['Texto']):
        lista2 = razao.loc[index, 'Montante em moeda interna']
        ppr2 += lista2
    elif row['Atribuição'] == '631.00.0.5' and 'conta' in row['Texto']:
        lista3 = razao.loc[index, 'Montante em moeda interna']
        outras_rec += lista3
    elif row['Atribuição'] == '631.00.0.5' and 'indébito' in row['Texto']:
        lista3a = razao.loc[index, 'Montante em moeda interna']
        indebito += lista3a
        print(indebito)
    elif row['Atribuição'] == '635.00.0.6' and 'conta' in row['Texto']:
        lista4 = razao.loc[index, 'Montante em moeda interna']
        outras_desp += lista4
    elif row['Atribuição'] == '710.00.2.1' and 'dif' in row['Texto']:
        lista5 = razao.loc[index, 'Montante em moeda interna']
        ir_dif += lista5
    elif row['Atribuição'] == '710.00.2.2' and 'dif' in row['Texto']:
        lista6 = razao.loc[index, 'Montante em moeda interna']
        csll_dif += lista6
    elif row['Atribuição'] == '675.03.0.3' and 'pis' in row['Texto']:
        lista_pis = razao.loc[index, 'Montante em moeda interna']
        prov_ind += lista_pis
    elif row['Atribuição'] == '675.03.0.3' and 'conta' in row['Texto']:
        lista_conta = razao.loc[index, 'Montante em moeda interna']
        deliberacao += lista_conta
    elif row['Atribuição'] == '675.03.0.3' and 'ajustek' in row['Texto']:
        lista_ajustek = razao.loc[index, 'Montante em moeda interna']
        termok += lista_ajustek
    elif row['Atribuição'] == '652.00.0.1.01' and 'vag' in row['Texto']:
        lista_vag = razao.loc[index, 'Montante em moeda interna']
        vag += lista_vag
    elif row['Atribuição'] == '651.00.0.1.01' and 'vag' in row['Texto']:
        lista_vag2 = razao.loc[index, 'Montante em moeda interna']
        vag2 += lista_vag2


# for row in ws.iter_rows(min_row=6, min_col=6, max_col=6, max_row=727):
#     for index, cell in enumerate(row):
#         if ws.cell(row=cell.row, column=6).value == ws.cell(row=cell.row - 1, column=6).value:
#             break
#         elif cell.value not in conciliacoes and cell.value not in razao:
#             for index, row2 in dados.iterrows():
#                 if str(cell.value) == row2['Conta CSPE']:
#                     ws.cell(row=cell.row, column=pos_coluna).value = dados.loc[index, ' Saldo Acumulado']
#                     ws.cell(row=cell.row, column=pos_coluna).number_format = '#,##0.00'
# =======================================================================================================#

# ====================================== DADOS BALANCETE PLANO 50 ======================================#
balancete = pd.read_excel(f'balancete{data_cur.month}.xlsx')
balancete = pd.DataFrame(balancete)

lista_plano50 = []
lista_pisco = []

# lista de contas 'diferença de PIS/COFINS'
contas_pisco = ['6110100004', '6112100004', '6113100004', '6114100004', '6116100004']

# Listar contas 615 (C/C, compensação) e dividir por segmento
contas_comp = [['6152131002', '6152131004'], ['6150131002', '6150131004', '6153131002', '6153131004'],
               ['6154131002', '6154131004'], ['6156131002', '6156131004']]

# dicionario para somar os valores nas contas chave
contas_custo = {'6152131002': 0, '6153131002': 0, '6154131002': 0, '6156131002': 0}

for index, row in balancete.iterrows():
    if row['Conta do Razão'] in contas_pisco:
        lista_pisco.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] in contas_comp[0]:
        contas_custo['6152131002'] += balancete.loc[index, 'Saldo Acumulado']
    elif row['Conta do Razão'] in contas_comp[1]:
        contas_custo['6153131002'] += balancete.loc[index, 'Saldo Acumulado']
    elif row['Conta do Razão'] in contas_comp[2]:
        contas_custo['6154131002'] += balancete.loc[index, 'Saldo Acumulado']
    elif row['Conta do Razão'] in contas_comp[3]:
        contas_custo['6156131002'] += balancete.loc[index, 'Saldo Acumulado']
    elif row['Conta do Razão'] == '1123120000':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '1123120001':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '6122102002':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '6123102002':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '6160191040':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '6160191041':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '6160351011':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '6310500002':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '6350600009':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])
    elif row['Conta do Razão'] == '6520101001':
        lista_plano50.append(balancete.loc[index, 'Saldo Acumulado'])

print(lista_plano50)
# =====================================================================================================#

dados = pd.read_excel(f'G:\GECOT\CONCILIAÇÕES CONTÁBEIS\CONCILIAÇÕES_2022\BALANCETES\SOCIETÁRIOS\\'
                      f'Balancete 0{data_cur.month}2022 rev.02.xlsx', skiprows=12)

dados = pd.DataFrame(dados)

dados['Conta CSPE'] = dados['Conta CSPE'].astype(str)

conciliacoes = ['121.51.9', '121.83.2', '211.21.1', '211.21.4', '211.61.1', '211.99.9', '221.61.1', '221.99.3',
                '223.03.1']
razao = []
caminho = f'G:\GECOT\CONCILIAÇÕES CONTÁBEIS\CONCILIAÇÕES_2022\\0{data_cur.month}.2022\\' + 'Conta '

# cont, cont2, cont3, cont4, cont5, cont6, cont7, cont8 = 0, 0, 0, 0, 0, 0, 0, 0
for row in ws.iter_rows(min_row=6, min_col=6, max_col=6, max_row=len(ws['F'])):
    for index, cell in enumerate(row):
        if ws.cell(row=cell.row, column=6).value == ws.cell(row=cell.row - 1, column=6).value:
            break
        elif cell.value not in conciliacoes and cell.value not in razao:
            for index2, row2 in dados.iterrows():
                if str(cell.value) == row2['Conta CSPE']:
                    ws.cell(row=cell.row, column=pos_coluna).value = dados.loc[index2, ' Saldo Acumulado']
                    ws.cell(row=cell.row, column=pos_coluna).number_format = '#,##0.00'

        elif cell.value == '121.51.9':
            wb2 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb2.sheetnames
            ws2 = wb2[sheets[0]]
            ws.cell(row=cell.row, column=pos_coluna).value = ws2['C5'].value
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws2['C5'].value
            ws.cell(row=cell.row + 2, column=pos_coluna).value = 0

        elif cell.value == '121.83.2':
            wb3 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb3.sheetnames
            ws3 = wb3[sheets[0]]
            ws.cell(row=cell.row, column=pos_coluna).value = ws3['C5'].value
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws3['I7'].value
            ws.cell(row=cell.row + 2, column=pos_coluna).value = 0
            ws.cell(row=cell.row + 3, column=pos_coluna).value = ws3['I8'].value
            ws.cell(row=cell.row + 4, column=pos_coluna).value = ws3['I6'].value

        elif cell.value == '211.21.1':
            wb4 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb4.sheetnames
            ws4 = wb4[sheets[0]]
            ws.cell(row=cell.row, column=pos_coluna).value = (ws4['D5'].value or 0) * -1
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws4['H4'].value or 0
            ws.cell(row=cell.row + 2, column=pos_coluna).value = (ws4['C7'].value or 0) * -1
            ws.cell(row=cell.row + 3, column=pos_coluna).value = ws4['H7'].value or 0
            ws.cell(row=cell.row + 4, column=pos_coluna).value = (ws4['H5'].value or 0) + (
                    (ws4['C8'].value or 0) * -1)
            ws.cell(row=cell.row + 5, column=pos_coluna).value = (ws4['H6'].value or 0) + (
                    (ws4['C9'].value or 0) * -1)
            ws.cell(row=cell.row + 6, column=pos_coluna).value = 0
            ws.cell(row=cell.row + 7, column=pos_coluna).value = ws4['H9'].value or 0

        elif cell.value == '211.21.4':
            wb4 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb4.sheetnames
            ws4 = wb4[sheets[0]]
            ws.cell(row=cell.row, column=pos_coluna).value = (ws4['D5'].value or 0) * - 1
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws4['H8'].value or 0
            ws.cell(row=cell.row + 2, column=pos_coluna).value = ws4['H7'].value or 0
            ws.cell(row=cell.row + 3, column=pos_coluna).value = ws4['H6'].value or 0
            ws.cell(row=cell.row + 4, column=pos_coluna).value = ws4['H9'].value or 0
            ws.cell(row=cell.row + 5, column=pos_coluna).value = ws4['H5'].value or 0
            ws.cell(row=cell.row + 6, column=pos_coluna).value = (ws4['H12'].value or 0) + (
                    (ws4['C7'].value or 0) * -1)
            ws.cell(row=cell.row + 7, column=pos_coluna).value = (ws4['H10'].value or 0) + ws4['H11'].value or 0

        elif cell.value == '211.61.1':
            wb2 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb2.sheetnames
            ws2 = wb2[sheets[0]]
            ws.cell(row=cell.row, column=pos_coluna).value = (ws2['C7'].value or 0) * -1
            ws.cell(row=cell.row + 1, column=pos_coluna).value = (ws2['C9'].value or 0) * -1
            ws.cell(row=cell.row + 2, column=pos_coluna).value = (ws2['C8'].value or 0) * -1

        elif cell.value == '221.61.1':
            wb2 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb2.sheetnames
            ws2 = wb2[sheets[0]]
            ws.cell(row=cell.row, column=pos_coluna).value = (ws2['D5'].value or 0) * -1
            ws.cell(row=cell.row + 1, column=pos_coluna).value = (ws2['C8'].value or 0) * -1
            ws.cell(row=cell.row + 2, column=pos_coluna).value = (ws2['C7'].value or 0) * -1

        elif cell.value == '221.99.3':
            wb2 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb2.sheetnames
            ws2 = wb2[sheets[0]]
            citygate, ambiental, pis, igas, multa, grafica = 0, 0, 0, 0, 0, 0
            for i in ws2['B']:
                if 'Multa' in str(i.value):
                    multa += ws2['C' + str(i.row)].value or 0
                elif 'AI.GAS' in str(i.value):
                    citygate += ws2['C' + str(i.row)].value or 0
                elif 'AMBIENTAL' in str(i.value):
                    ambiental += ws2['C' + str(i.row)].value or 0
                elif 'PIS/COFINS' in str(i.value):
                    pis += ws2['C' + str(i.row)].value or 0
                elif 'Igás' in str(i.value):
                    igas += ws2['C' + str(i.row)].value or 0
                elif 'GRÁFICA' in str(i.value):
                    grafica += ws2['C' + str(i.row)].value or 0

            ws.cell(row=cell.row, column=pos_coluna).value = (ws2['D5'].value or 0) * -1
            ws.cell(row=cell.row + 1, column=pos_coluna).value = 0
            ws.cell(row=cell.row + 2, column=pos_coluna).value = 0
            # ws.cell(row=cell.row + 2, column=pos_coluna).value = citygate * -1
            ws.cell(row=cell.row + 3, column=pos_coluna).value = (ws2['I8'].value or 0) * -1
            # ws.cell(row=cell.row + 3, column=pos_coluna).value = multa * -1
            ws.cell(row=cell.row + 4, column=pos_coluna).value = (ws2['I9'].value or 0) * -1
            # ws.cell(row=cell.row + 4, column=pos_coluna).value = ambiental * -1
            ws.cell(row=cell.row + 5, column=pos_coluna).value = 0
            # ws.cell(row=cell.row + 6, column=pos_coluna).value = pis * -1
            ws.cell(row=cell.row + 6, column=pos_coluna).value = (ws2['I11'].value or 0) * -1
            ws.cell(row=cell.row + 7, column=pos_coluna).value = 0
            # ws.cell(row=cell.row + 8, column=pos_coluna).value = igas * -1
            ws.cell(row=cell.row + 8, column=pos_coluna).value = (ws2['I12'].value or 0) * -1
            ws.cell(row=cell.row + 9, column=pos_coluna).value = (ws2['I10'].value or 0) * -1

        elif cell.value == '211.99.9':
            wb4 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb4.sheetnames
            ws4 = wb4[sheets[0]]
            ws.cell(row=cell.row, column=pos_coluna).value = (ws4['D5'].value or 0) * -1
            ws.cell(row=cell.row + 1, column=pos_coluna).value = (ws4['G10'].value or 0) * -1
            ws.cell(row=cell.row + 2, column=pos_coluna).value = (ws4['G11'].value or 0) * -1
            ws.cell(row=cell.row + 3, column=pos_coluna).value = (ws4['G9'].value or 0) * -1
            ws.cell(row=cell.row + 4, column=pos_coluna).value = (ws4['H6'].value or 0) * -1
            ws.cell(row=cell.row + 5, column=pos_coluna).value = (ws4['G14'].value or 0) * -1
            ws.cell(row=cell.row + 6, column=pos_coluna).value = 0
            ws.cell(row=cell.row + 7, column=pos_coluna).value = (ws4['G13'].value or 0) * -1
            ws.cell(row=cell.row + 8, column=pos_coluna).value = 0

        elif cell.value == '223.03.1':
            wb2 = openpyxl.load_workbook(caminho + cell.value.replace('.', '') + '.xlsx', data_only=True)
            sheets = wb2.sheetnames
            ws2 = wb2[sheets[0]]
            ws.cell(row=cell.row, column=pos_coluna).value = ws2['D5'].value * -1
            ws.cell(row=cell.row+1, column=pos_coluna).value = ws2['C7'].value * -1
            ws.cell(row=cell.row+2, column=pos_coluna).value = ws2['C8'].value * -1

        if cell.value == '615.03.1.1.01.01':
            ws.cell(row=cell.row, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - ppr
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ppr
        if cell.value == '616.00.0.1.01.01':
            ws.cell(row=cell.row, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - ppr1
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ppr1
        if cell.value == '616.00.0.3.01.01':
            ws.cell(row=cell.row, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - ppr2
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ppr2
        if cell.value == '710.00.2.1':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row,
                                                                         column=pos_coluna).value - ir_dif
            ws.cell(row=cell.row + 2, column=pos_coluna).value = ir_dif
        if cell.value == '710.00.2.2':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row,
                                                                         column=pos_coluna).value - csll_dif
            ws.cell(row=cell.row + 2, column=pos_coluna).value = csll_dif
        # valores balancete plano 50
        if cell.value == '611.03.2.2':
            ws.cell(row=cell.row, column=pos_coluna).value = \
                ws.cell(row=cell.row, column=pos_coluna).value + lista_pisco[1]
            ws.cell(row=cell.row + 1, column=pos_coluna).value = lista_pisco[1] * -1
        if cell.value == '611.03.3.2':
            ws.cell(row=cell.row, column=pos_coluna).value = \
                ws.cell(row=cell.row, column=pos_coluna).value + lista_pisco[2] + lista_pisco[0]
            ws.cell(row=cell.row + 1, column=pos_coluna).value = (lista_pisco[2] + lista_pisco[0]) * -1
        if cell.value == '611.03.4.2':
            ws.cell(row=cell.row, column=pos_coluna).value = sum(filter(None,
                [ws.cell(row=cell.row, column=pos_coluna).value, lista_pisco[3]]))
            ws.cell(row=cell.row + 1, column=pos_coluna).value = lista_pisco[3] * -1
        if cell.value == '611.03.6.1':
            ws.cell(row=cell.row, column=pos_coluna).value = \
                ws.cell(row=cell.row, column=pos_coluna).value + lista_pisco[4]
            ws.cell(row=cell.row + 1, column=pos_coluna).value = lista_pisco[4] * -1
        if cell.value == '615.03.2.1.31':
            ws.cell(row=cell.row, column=pos_coluna).value = \
                ws.cell(row=cell.row, column=pos_coluna).value - contas_custo['6152131002']
            ws.cell(row=cell.row + 1, column=pos_coluna).value = contas_custo['6152131002']
            ws.cell(row=cell.row + 2, column=pos_coluna).value = 0
        if cell.value == '615.03.3.1.31':
            ws.cell(row=cell.row, column=pos_coluna).value = \
                ws.cell(row=cell.row, column=pos_coluna).value - contas_custo['6153131002']
            ws.cell(row=cell.row + 1, column=pos_coluna).value = contas_custo['6153131002']
            ws.cell(row=cell.row + 2, column=pos_coluna).value = 0
        if cell.value == '615.03.4.1.31':
            ws.cell(row=cell.row, column=pos_coluna).value = \
                ws.cell(row=cell.row, column=pos_coluna).value - contas_custo['6154131002']
            ws.cell(row=cell.row + 1, column=pos_coluna).value = contas_custo['6154131002']
            ws.cell(row=cell.row + 2, column=pos_coluna).value = 0
        if cell.value == '615.03.6.1.31':
            ws.cell(row=cell.row, column=pos_coluna).value = \
                ws.cell(row=cell.row, column=pos_coluna).value - contas_custo['6156131002']
            ws.cell(row=cell.row + 1, column=pos_coluna).value = contas_custo['6156131002']
            ws.cell(row=cell.row + 2, column=pos_coluna).value = 0
        if cell.value == '112.31.2':
            ws.cell(row=cell.row, column=pos_coluna).value = 0
            ws.cell(row=cell.row + 1, column=pos_coluna).value = lista_plano50[0]
        if cell.value == '612.03.2.1.02':
            ws.cell(row=cell.row, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - \
                                                             lista_plano50[1]
            ws.cell(row=cell.row + 1, column=pos_coluna).value = lista_plano50[1]
        if cell.value == '612.03.3.1.02':
            ws.cell(row=cell.row, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - \
                                                             lista_plano50[2]
            ws.cell(row=cell.row + 1, column=pos_coluna).value = lista_plano50[2]
        if cell.value == '616.00.0.1.91.04':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = lista_plano50[3]
            ws.cell(row=cell.row + 2, column=pos_coluna).value = lista_plano50[4]
        if cell.value == '616.00.0.3.51.01':
            ws.cell(row=cell.row, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - \
                                                             lista_plano50[5]
            ws.cell(row=cell.row + 1, column=pos_coluna).value = lista_plano50[5]
        if cell.value == '631.00.0.5':
            ws.cell(row=cell.row, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - \
                                                             indebito - outras_rec - lista_plano50[6]
            ws.cell(row=cell.row + 1, column=pos_coluna).value = outras_rec
            ws.cell(row=cell.row + 2, column=pos_coluna).value = lista_plano50[6]
            ws.cell(row=cell.row + 3, column=pos_coluna).value = indebito
        if cell.value == '635.00.0.6':
            ws.cell(row=cell.row, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - \
                                                                 outras_desp - (lista_plano50[7])
            ws.cell(row=cell.row + 2, column=pos_coluna).value = outras_desp
            ws.cell(row=cell.row + 3, column=pos_coluna).value = lista_plano50[7]
        if cell.value == '616.00.0.3.42.01':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value
        if cell.value == '631.00.0.2':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value
        if cell.value == '635.00.0.2':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value
        if cell.value == '651.00.0.1.01':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = vag2
            ws.cell(row=cell.row + 2, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value - vag2
        if cell.value == '652.00.0.1.01':
            ws.cell(row=cell.row + 2, column=pos_coluna).value = lista_plano50[8]
            ws.cell(row=cell.row + 3, column=pos_coluna).value = vag
        if cell.value == '671.03.0.3':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value
        if cell.value == '675.03.0.3':
            ws.cell(row=cell.row, column=pos_coluna).value = 0
            ws.cell(row=cell.row+1, column=pos_coluna).value = prov_ind
            ws.cell(row=cell.row+2, column=pos_coluna).value = deliberacao
            ws.cell(row=cell.row+3, column=pos_coluna).value = termok
        if cell.value == '616.00.0.3.94':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value
        if cell.value == '616.00.0.3.94.01':
            ws.cell(row=cell.row + 1, column=pos_coluna).value = ws.cell(row=cell.row, column=pos_coluna).value
            ws.cell(row=cell.row, column=pos_coluna).value = 0

wb.save('DF2.xlsx')
